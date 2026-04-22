import io
import re
import os
from flask import Flask, request, send_file, render_template, jsonify
import openpyxl
from openpyxl import load_workbook

app = Flask(__name__)

DOMAIN_MAP = {
    "strongmindsandstrongbodies": "D1",
    "positiveselfworth": "D2",
    "caringfamiliesandrelationships": "D3",
    "safety": "D4",
    "vibrantcommunities": "D5",
    "healthyenvironments": "D6",
    "racialjusticeequityandinclusion": "D7",
    "funnandhappiness": "D8",
    "funandhappiness": "D8",
}

def normalize(s):
    return re.sub(r'[^a-z0-9]', '', s.lower())

def resolve_domain(user_input):
    key = normalize(user_input)
    return DOMAIN_MAP.get(key)

def extract_q_label(col_name):
    """Extract 'Q1', 'Q2', etc. from a column name like 'Q1. Please rate...'"""
    if col_name is None:
        return None
    m = re.match(r'(Q\d+)', str(col_name), re.IGNORECASE)
    return m.group(1).upper() if m else None

def convert_voters_sheet(file_bytes, domain_code):
    wb = load_workbook(filename=io.BytesIO(file_bytes))
    if 'Voters' not in wb.sheetnames:
        raise ValueError("No 'Voters' sheet found in the workbook.")
    ws = wb['Voters']

    # Find header row: first row where first cell is 'Date'
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0] and str(row[0]).strip().lower() == 'date':
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("Could not find header row with 'Date' column.")

    headers = [cell.value for cell in ws[header_row_idx]]

    # Identify columns to keep and their new names
    # Delete: Date (col index 0), Session (col index 1), col D (col index 3)
    # Keep: Voter (col index 2), then Q columns (index 4+)
    # Col indices are 0-based here

    new_headers = []
    keep_col_indices = []

    for i, h in enumerate(headers):
        if i in (0, 1, 3):  # Date, Session, col D
            continue
        if i == 2:  # Voter
            new_headers.append('Voter')
            keep_col_indices.append(i)
        else:
            q_label = extract_q_label(h)
            if q_label:
                new_headers.append(f"{domain_code}_{q_label}")
                keep_col_indices.append(i)
            # skip columns that don't match Q pattern

    # Build output workbook
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Voters'

    out_ws.append(new_headers)

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue
        out_row = [row[i] for i in keep_col_indices]
        out_ws.append(out_row)

    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return buf

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/convert', methods=['POST'])
def convert():
    if 'file' not in request.files or request.files['file'].filename == '':
        return jsonify({'error': 'No file uploaded.'}), 400
    domain_input = request.form.get('domain', '').strip()
    if not domain_input:
        return jsonify({'error': 'Please enter a domain name.'}), 400

    domain_code = resolve_domain(domain_input)
    if not domain_code:
        return jsonify({'error': f'Unrecognized domain: "{domain_input}". Please check your spelling.'}), 400

    file = request.files['file']
    file_bytes = file.read()
    try:
        result_buf = convert_voters_sheet(file_bytes, domain_code)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    return send_file(
        result_buf,
        as_attachment=True,
        download_name='Voters_converted.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5050))
    app.run(debug=False, host='0.0.0.0', port=port)
